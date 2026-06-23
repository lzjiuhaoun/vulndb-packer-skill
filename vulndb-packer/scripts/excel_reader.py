"""
Excel读取器模块
集成excel-reader skill的核心功能
使用Windows COM接口或openpyxl读取Excel文件
"""

import os
import re
import sys
from typing import Dict, List, Any, Optional

# 尝试导入COM支持
try:
    import pythoncom
    import win32com.client
    HAS_COM_SUPPORT = True
except ImportError as e:
    HAS_COM_SUPPORT = False

# 尝试导入openpyxl
try:
    import openpyxl
    HAS_OPENPYXL_SUPPORT = True
except ImportError as e:
    HAS_OPENPYXL_SUPPORT = False


class ExcelReadError(Exception):
    """Excel文档读取错误"""
    def __init__(self, file_path: str, reason: str):
        self.file_path = file_path
        self.reason = reason
        super().__init__(f"无法读取Excel文件 [{file_path}]: {reason}")


class ExcelReader:
    """Excel读取器
    
    通过Windows COM接口或openpyxl打开Excel文档，
    提取所有工作表内容并转换为Markdown表格。
    """
    
    def __init__(self):
        """初始化Excel读取器"""
        self.excel_app = None
        self.use_com = False
        
        # 优先使用COM接口
        if HAS_COM_SUPPORT:
            try:
                self._init_excel_app()
                self.use_com = True
            except Exception as e:
                print(f"警告: COM接口初始化失败，将使用openpyxl: {e}")
        
        # 如果COM不可用，使用openpyxl
        if not self.use_com:
            if not HAS_OPENPYXL_SUPPORT:
                raise ExcelReadError("", "需要安装pywin32或openpyxl库: pip install pywin32 或 pip install openpyxl")
    
    def _init_excel_app(self):
        """初始化Excel.Application COM对象"""
        try:
            pythoncom.CoInitialize()
            self.excel_app = win32com.client.Dispatch("Excel.Application")
            self.excel_app.Visible = False
            try:
                self.excel_app.DisplayAlerts = False
            except Exception:
                pass
        except Exception as e:
            raise ExcelReadError("", f"Excel COM初始化失败: {str(e)}")
    
    def read_excel(self, file_path: str) -> Dict[str, Any]:
        """
        读取单个Excel文件
        
        Args:
            file_path: Excel文件绝对路径
            
        Returns:
            {
                "file_name": "工作簿名称",
                "sheet_count": 3,
                "sheets": [
                    {"name": "Sheet1", "markdown": "| ... |", "row_count": 10, "col_count": 5},
                    ...
                ]
            }
            
        Raises:
            ExcelReadError: 文件不存在、格式不支持、读取失败等
        """
        # 验证文件存在
        if not os.path.exists(file_path):
            raise ExcelReadError(file_path, "文件路径不存在")
        
        # 验证文件扩展名
        if not file_path.lower().endswith((".xls", ".xlsx")):
            raise ExcelReadError(file_path, "不支持的文件格式，仅支持 .xls/.xlsx")
        
        # 使用COM接口读取
        if self.use_com:
            return self._read_excel_com(file_path)
        else:
            return self._read_excel_openpyxl(file_path)
    
    def _read_excel_com(self, file_path: str) -> Dict[str, Any]:
        """使用COM接口读取Excel文件"""
        workbook = None
        try:
            workbook = self.excel_app.Workbooks.Open(
                file_path,
                ReadOnly=True,
                UpdateLinks=0,
            )
            
            file_name = os.path.basename(file_path)
            sheets_data = []
            
            for sheet in workbook.Worksheets:
                sheet_data = self._read_sheet_com(sheet)
                sheets_data.append(sheet_data)
            
            return {
                "file_name": file_name,
                "sheet_count": len(sheets_data),
                "sheets": sheets_data,
            }
            
        except Exception as e:
            raise ExcelReadError(file_path, f"COM调用失败: {str(e)}")
        finally:
            if workbook is not None:
                try:
                    workbook.Close(SaveChanges=False)
                except Exception:
                    pass
    
    def _read_excel_openpyxl(self, file_path: str) -> Dict[str, Any]:
        """使用openpyxl读取Excel文件"""
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            file_name = os.path.basename(file_path)
            sheets_data = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = self._read_sheet_openpyxl(sheet, sheet_name)
                sheets_data.append(sheet_data)
            
            workbook.close()
            
            return {
                "file_name": file_name,
                "sheet_count": len(sheets_data),
                "sheets": sheets_data,
            }
            
        except Exception as e:
            raise ExcelReadError(file_path, f"openpyxl读取失败: {str(e)}")
    
    def read_multiple_excels(self, file_paths: List[str]) -> Dict[str, Any]:
        """
        批量读取多个Excel文件
        
        Args:
            file_paths: Excel文件路径列表
            
        Returns:
            {
                "success": [
                    {"file": "路径1", "data": {...}},
                    ...
                ],
                "failed": [
                    {"path": "路径3", "error": "错误描述"}
                ]
            }
        """
        success_list = []
        failed_list = []
        
        for file_path in file_paths:
            try:
                result = self.read_excel(file_path)
                success_list.append({"file": file_path, "data": result})
            except ExcelReadError as e:
                failed_list.append({"path": file_path, "error": e.reason})
            except Exception as e:
                failed_list.append({"path": file_path, "error": str(e)})
        
        return {"success": success_list, "failed": failed_list}
    
    def _read_sheet_com(self, sheet) -> Dict[str, Any]:
        """使用COM接口读取单个工作表并转换为Markdown表格"""
        try:
            used_range = sheet.UsedRange
            if used_range is None:
                return {
                    "name": sheet.Name,
                    "markdown": "*（空工作表）*",
                    "row_count": 0,
                    "col_count": 0,
                }
            
            row_count = used_range.Rows.Count
            col_count = used_range.Columns.Count
            
            rows = []
            for i in range(1, row_count + 1):
                row_values = []
                for j in range(1, col_count + 1):
                    cell = used_range.Cells(i, j)
                    value = cell.Value if cell.Value is not None else ""
                    value_str = str(value) if value else ""
                    value_str = self._escape_markdown(value_str)
                    value_str = value_str.replace("\n", "<br>").replace("\r", "")
                    row_values.append(value_str)
                rows.append(row_values)
            
            markdown = self._to_markdown_table(rows)
            
            return {
                "name": sheet.Name,
                "markdown": markdown,
                "row_count": row_count,
                "col_count": col_count,
            }
            
        except Exception as e:
            return {
                "name": sheet.Name,
                "markdown": f"*（读取工作表失败: {str(e)}）*",
                "row_count": 0,
                "col_count": 0,
            }
    
    def _read_sheet_openpyxl(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """使用openpyxl读取单个工作表并转换为Markdown表格"""
        try:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                row_values = []
                for value in row:
                    value_str = str(value) if value is not None else ""
                    value_str = self._escape_markdown(value_str)
                    value_str = value_str.replace("\n", "<br>").replace("\r", "")
                    row_values.append(value_str)
                rows.append(row_values)
            
            if not rows:
                return {
                    "name": sheet_name,
                    "markdown": "*（空工作表）*",
                    "row_count": 0,
                    "col_count": 0,
                }
            
            row_count = len(rows)
            col_count = len(rows[0]) if rows else 0
            
            markdown = self._to_markdown_table(rows)
            
            return {
                "name": sheet_name,
                "markdown": markdown,
                "row_count": row_count,
                "col_count": col_count,
            }
            
        except Exception as e:
            return {
                "name": sheet_name,
                "markdown": f"*（读取工作表失败: {str(e)}）*",
                "row_count": 0,
                "col_count": 0,
            }
    
    def _escape_markdown(self, text: str) -> str:
        """转义Markdown表格中的特殊字符"""
        text = text.replace("|", "&#124;")
        return text
    
    def _to_markdown_table(self, rows: List[List[str]]) -> str:
        """将二维数组转换为Markdown表格"""
        if not rows or not rows[0]:
            return "*（空表格）*"
        
        col_count = len(rows[0])
        col_widths = [0] * col_count
        
        for row in rows:
            for j, cell in enumerate(row):
                if j < col_count:
                    col_widths[j] = max(col_widths[j], len(cell))
        
        lines = []
        
        header_cells = [rows[0][j].ljust(col_widths[j]) for j in range(col_count)]
        lines.append("| " + " | ".join(header_cells) + " |")
        
        separator_cells = ["-" * (col_widths[j] + 2) for j in range(col_count)]
        lines.append("|" + "|".join(separator_cells) + "|")
        
        for row in rows[1:]:
            row_cells = [row[j].ljust(col_widths[j]) if j < len(row) else "".ljust(col_widths[j]) for j in range(col_count)]
            lines.append("| " + " | ".join(row_cells) + " |")
        
        return "\n".join(lines)
    
    def close(self):
        """关闭Excel应用，释放资源"""
        if self.excel_app is not None:
            try:
                self.excel_app.Quit()
            except Exception:
                pass
        try:
            if HAS_COM_SUPPORT:
                pythoncom.CoUninitialize()
        except Exception:
            pass
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
        return False
